import streamlit as st
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import os
import re
import time
import mimetypes
import base64

st.set_page_config(page_title="Excel 图片处理工具", page_icon="📊", layout="wide")
st.title("📊 Excel 图片处理工具")
st.write("上传包含图片链接的 Excel 文件，自动下载图片并插入表格，支持 webp 转 png。")

# --- 工具函数 ---
def register_webp_mimetype():
    try:
        if '.webp' not in mimetypes.types_map:
            mimetypes.add_type('image/webp', '.webp')
        st.success("✅ 已注册 .webp 格式")
    except Exception as e:
        st.warning(f"⚠️ MIME注册警告: {str(e)}")

def convert_webp_to_png(webp_path):
    try:
        png_path = os.path.splitext(webp_path)[0] + '.png'
        with PILImage.open(webp_path) as img:
            if img.mode in ('RGBA', 'LA'):
                background = PILImage.new(img.mode[:-1], img.size, (255, 255, 255))
                background.paste(img, img.split()[-1])
                background.save(png_path, 'PNG')
            else:
                img.save(png_path, 'PNG')
        return png_path
    except:
        return None

def create_temp_folders():
    for folder in ['temp_images', 'output', 'temp_png']:
        if not os.path.exists(folder):
            os.makedirs(folder)

def is_image_url(url):
    if not url or not isinstance(url, str):
        return False
    if not url.startswith(('http://','https://')):
        return False
    return any(ext in url.lower() for ext in ['webp','jpg','jpeg','png','gif','bmp','svg'])

def download_and_convert_image(image_url, save_folder, convert_folder):
    try:
        image_format = 'jpg'
        is_webp = 'webp' in image_url.lower()
        if is_webp:
            image_format = 'webp'
        elif 'png' in image_url.lower():
            image_format = 'png'
        elif any(ext in image_url.lower() for ext in ['jpg','jpeg']):
            image_format = 'jpg'

        original_name = image_url.split("/")[-1]
        if not os.path.splitext(original_name)[1]:
            original_name += f".{image_format}"
        safe_name = re.sub(r'[^\w\.]', '_', original_name)
        local_img_path = os.path.join(save_folder, safe_name)

        if os.path.exists(local_img_path):
            if is_webp:
                png_path = os.path.join(convert_folder, os.path.splitext(safe_name)[0]+'.png')
                if os.path.exists(png_path):
                    return png_path
            else:
                return local_img_path

        headers = {"User-Agent":"Mozilla/5.0"}
        response = requests.get(image_url, headers=headers, timeout=20)
        response.raise_for_status()
        with open(local_img_path,'wb') as f:
            f.write(response.content)

        if is_webp:
            png_path = convert_webp_to_png(local_img_path)
            if png_path:
                new_png_path = os.path.join(convert_folder, os.path.basename(png_path))
                os.rename(png_path, new_png_path)
                return new_png_path
            else:
                return local_img_path
        else:
            return local_img_path
    except:
        return None

def auto_insert_images_to_excel(excel_path, sheet_name, st_container=None):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    image_links = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and is_image_url(cell.value.strip()):
                image_links.append({'url': cell.value.strip(), 'row': cell.row, 'column': cell.column_letter, 'cell': f"{cell.column_letter}{cell.row}"})

    total = len(image_links)
    progress_bar = st_container.progress(0)
    status_text = st_container.empty()
    start_time = time.time()
    success_count = 0
    fail_count = 0

    for i, link_info in enumerate(image_links, 1):
        status_text.text(f"处理第 {i}/{total} 个图片: {link_info['cell']}，成功 {success_count} 张，失败 {fail_count} 张")
        local_img_path = download_and_convert_image(link_info['url'], 'temp_images','temp_png')
        if not local_img_path:
            fail_count += 1
            progress_bar.progress(int(i/total*100))
            continue
        ws.row_dimensions[link_info['row']].height = 120
        ws.column_dimensions[link_info['column']].width = 25
        try:
            img = Image(local_img_path)
            max_width = ws.column_dimensions[link_info['column']].width*6
            max_height = ws.row_dimensions[link_info['row']].height*1.33
            if img.width>max_width or img.height>max_height:
                scale_ratio = min(max_width/img.width, max_height/img.height)
                img.width = int(img.width*scale_ratio)
                img.height = int(img.height*scale_ratio)
            ws.add_image(img, link_info['cell'])
            success_count += 1
        except:
            fail_count += 1

        # 更新进度条和预估时间
        progress_bar.progress(int(i/total*100))
        elapsed = time.time() - start_time
        avg_per_item = elapsed / i
        remaining = int(avg_per_item * (total - i))
        status_text.text(
            f"处理第 {i}/{total} 个图片: {link_info['cell']}\n"
            f"成功 {success_count} 张，失败 {fail_count} 张\n"
            f"预计剩余时间: {remaining} 秒"
        )
    output_path = os.path.join('output', f"带图片_{os.path.basename(excel_path)}")
    wb.save(output_path)
    return output_path, success_count, fail_count

# --- Streamlit 页面 ---
register_webp_mimetype()
create_temp_folders()

uploaded_file = st.file_uploader("📁 上传 Excel 文件 (.xlsx)", type=["xlsx"])
sheet_name = st.text_input("工作表名称（默认Sheet1）", value="Sheet1")

if uploaded_file and st.button("开始处理"):
    with open(os.path.join("temp_images", uploaded_file.name), "wb") as f:
        f.write(uploaded_file.getbuffer())
    st_container = st.empty()
    st.info("⏳ 开始处理，请稍候...")
    output_path, success_count, fail_count = auto_insert_images_to_excel(
        os.path.join("temp_images", uploaded_file.name),
        sheet_name,
        st_container
    )
    st.success(f"✅ 处理完成！成功插入 {success_count} 张图片，失败 {fail_count} 张")

    # 下载按钮
    with open(output_path, "rb") as f:
        bytes_data = f.read()
        b64 = base64.b64encode(bytes_data).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{os.path.basename(output_path)}">📥 下载处理后的 Excel 文件</a>'
        st.markdown(href, unsafe_allow_html=True)
